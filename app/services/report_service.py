"""Report generation service using ReportLab for PDF generation."""
from pathlib import Path
from typing import Dict, Any
from io import BytesIO
import openpyxl
from docxtpl import DocxTemplate
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


class ReportService:
    """Service for generating reports from templates."""

    def __init__(self):
        self.template_dir = Path(__file__).parent.parent / "templates" / "reports"

    async def generate_order_confirmation_pdf(
        self,
        data: Dict[str, Any]
    ) -> bytes:
        """
        Generate order confirmation PDF using ReportLab.

        Args:
            data: Dictionary with order data

        Returns:
            PDF file as bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=30,
            alignment=1,  # Center
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=12,
            borderWidth=2,
            borderColor=colors.HexColor('#2c5282'),
            borderPadding=6,
        )

        normal_style = styles['Normal']
        normal_style.fontSize = 11

        # Header
        elements.append(Paragraph("ORDREBEKREFTELSE", title_style))
        elements.append(Paragraph("Larvik Kommune Catering", styles['Normal']))
        elements.append(Spacer(1, 2*cm))

        # Order details section
        elements.append(Paragraph("Ordredetaljer", heading_style))
        order_details = [
            ["Ordrenummer:", str(data.get('ordrenummer', ''))],
            ["Ordredato:", data.get('ordredato', '')],
            ["Leveringsdato:", data.get('leveringsdato', '')]
        ]
        order_table = Table(order_details, colWidths=[4*cm, 12*cm])
        order_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(order_table)
        elements.append(Spacer(1, 1*cm))

        # Customer section
        elements.append(Paragraph("Kunde", heading_style))
        kunde = data.get('kunde', {})
        customer_details = [
            ["Navn:", kunde.get('navn', '')],
            ["Adresse:", kunde.get('adresse', '')],
            ["Postnr/Sted:", f"{kunde.get('postnr', '')} {kunde.get('sted', '')}"]
        ]
        customer_table = Table(customer_details, colWidths=[4*cm, 12*cm])
        customer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(customer_table)
        elements.append(Spacer(1, 1*cm))

        # Products section
        elements.append(Paragraph("Produkter", heading_style))

        # Products table
        products_data = [["Produktnavn", "Antall", "Pris", "Sum"]]
        for produkt in data.get('produkter', []):
            products_data.append([
                produkt.get('navn', ''),
                f"{produkt.get('antall', '')} {produkt.get('enhet', '')}",
                f"kr {produkt.get('pris', '')}",
                f"kr {produkt.get('sum', '')}"
            ])

        # Total row
        products_data.append([
            '', '', 'TOTAL:', f"kr {data.get('totalsum', '')}"
        ])

        products_table = Table(products_data, colWidths=[8*cm, 3*cm, 2.5*cm, 2.5*cm])
        products_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),

            # Data rows
            ('FONTSIZE', (0, 1), (-1, -2), 11),
            ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9f9f9')]),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),

            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e6f2ff')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 13),
            ('ALIGN', (2, -1), (-1, -1), 'RIGHT'),

            # All rows
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(products_table)

        # Additional information if exists
        if data.get('informasjon'):
            elements.append(Spacer(1, 1*cm))
            info_style = ParagraphStyle(
                'InfoBox',
                parent=styles['Normal'],
                fontSize=11,
                leftIndent=10,
                rightIndent=10,
                spaceBefore=10,
                spaceAfter=10,
                backColor=colors.HexColor('#fff9e6'),
                borderWidth=1,
                borderColor=colors.HexColor('#f0ad4e'),
                borderPadding=10,
            )
            elements.append(Paragraph(f"<b>Tilleggsinformasjon:</b><br/>{data.get('informasjon')}", info_style))

        # Footer
        elements.append(Spacer(1, 2*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#666666'),
        )
        elements.append(Paragraph(f"Generert: {data.get('generert_dato', '')}", footer_style))
        elements.append(Paragraph("Larvik Kommune Catering - Telefon: XXX XX XXX - E-post: kontakt@lkc.no", footer_style))

        # Build PDF
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    async def generate_delivery_note_pdf(
        self,
        data: Dict[str, Any]
    ) -> bytes:
        """
        Generate delivery note PDF using ReportLab.

        Args:
            data: Dictionary with delivery data

        Returns:
            PDF file as bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=30,
            alignment=1,
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=12,
            borderWidth=2,
            borderColor=colors.HexColor('#2c5282'),
            borderPadding=6,
        )

        # Header
        elements.append(Paragraph("LEVERINGSSEDDEL", title_style))
        elements.append(Paragraph("Larvik Kommune Catering", styles['Normal']))
        elements.append(Spacer(1, 2*cm))

        # Delivery details
        elements.append(Paragraph("Leveringsdetaljer", heading_style))
        delivery_details = [
            ["Ordrenummer:", str(data.get('ordrenummer', ''))],
            ["Leveringsdato:", data.get('leveringsdato', '')]
        ]
        delivery_table = Table(delivery_details, colWidths=[4*cm, 12*cm])
        delivery_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(delivery_table)
        elements.append(Spacer(1, 1*cm))

        # Recipient section
        elements.append(Paragraph("Mottaker", heading_style))
        kunde = data.get('kunde', {})
        recipient_details = [
            ["Navn:", kunde.get('navn', '')],
            ["Adresse:", kunde.get('adresse', '')],
            ["Postnr/Sted:", f"{kunde.get('postnr', '')} {kunde.get('sted', '')}"]
        ]
        recipient_table = Table(recipient_details, colWidths=[4*cm, 12*cm])
        recipient_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(recipient_table)
        elements.append(Spacer(1, 1*cm))

        # Products section
        elements.append(Paragraph("Produkter", heading_style))

        products_data = [["Produktnavn", "Antall", "Mottatt"]]
        for produkt in data.get('produkter', []):
            products_data.append([
                produkt.get('navn', ''),
                f"{produkt.get('antall', '')} {produkt.get('enhet', '')}",
                ''  # Empty checkbox column
            ])

        products_table = Table(products_data, colWidths=[10*cm, 3*cm, 3*cm])
        products_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),

            # Data rows
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # Checkbox column
            ('BOX', (2, 1), (2, -1), 1, colors.HexColor('#999999')),

            # All rows
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(products_table)

        # Signature section
        elements.append(Spacer(1, 3*cm))
        elements.append(Paragraph("Signatur", heading_style))
        elements.append(Spacer(1, 1*cm))

        signature_data = [
            [Paragraph("<b>Utlevert av:</b>", styles['Normal']),
             Paragraph("<b>Mottatt av:</b>", styles['Normal'])],
            ['', ''],  # Space for signature
            ['_' * 40, '_' * 40],  # Signature line
            [Paragraph("<font size=9>Navn og dato</font>", styles['Normal']),
             Paragraph("<font size=9>Navn og dato</font>", styles['Normal'])]
        ]
        signature_table = Table(signature_data, colWidths=[8*cm, 8*cm], rowHeights=[None, 2*cm, None, None])
        signature_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(signature_table)

        # Footer
        elements.append(Spacer(1, 2*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#666666'),
        )
        elements.append(Paragraph(f"Generert: {data.get('generert_dato', '')}", footer_style))
        elements.append(Paragraph("Larvik Kommune Catering - Telefon: XXX XX XXX - E-post: kontakt@lkc.no", footer_style))

        # Build PDF
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    async def generate_pdf_from_docx(
        self,
        template_name: str,
        data: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF from Word template.

        Args:
            template_name: Name of .docx template file
            data: Dictionary with data to inject into template

        Returns:
            PDF file as bytes
        """
        template_path = self.template_dir / template_name

        # Load docx template
        doc = DocxTemplate(str(template_path))

        # Render with data
        doc.render(data)

        # Save to BytesIO
        docx_bytes = BytesIO()
        doc.save(docx_bytes)
        docx_bytes.seek(0)

        return docx_bytes.getvalue()

    async def generate_excel(
        self,
        template_name: str,
        data: Dict[str, Any]
    ) -> bytes:
        """
        Generate Excel file from template.

        Args:
            template_name: Name of .xlsx template file
            data: Dictionary with data to populate Excel

        Returns:
            Excel file as bytes
        """
        template_path = self.template_dir / template_name

        # Load template
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active

        # For simple data lists (e.g., customer list)
        if "rows" in data:
            start_row = data.get("start_row", 2)  # Default to row 2 (after header)

            for i, row_data in enumerate(data["rows"], start=start_row):
                for col_idx, value in enumerate(row_data.values(), start=1):
                    ws.cell(row=i, column=col_idx, value=value)

        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        return excel_bytes.getvalue()

    async def generate_simple_excel(
        self,
        headers: list[str],
        rows: list[list[Any]],
        sheet_name: str = "Sheet1"
    ) -> bytes:
        """
        Generate Excel file from scratch (no template).

        Args:
            headers: Column headers
            rows: Data rows
            sheet_name: Name of the Excel sheet

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        return excel_bytes.getvalue()
